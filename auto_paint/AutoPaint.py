import openpyxl
from openpyxl.utils import get_column_letter
import os

class AutoPaint:
    def __init__(self, filepath):
        self.filepath = filepath
        self.workbook = openpyxl.load_workbook(filepath)
        self.worksheet = self.workbook.active

    def adicionar_colunas_status_funcao(self):
        max_col = self.worksheet.max_column
        max_row = self.worksheet.max_row

        self.status_col = max_col + 1
        self.worksheet.cell(row=1, column=self.status_col).value = "Status"

        self.formula_col = self.status_col + 1
        formula_letter = get_column_letter(self.formula_col)
        self.worksheet.cell(row=1, column=self.formula_col).value = "Função"

        status_letter = get_column_letter(self.status_col)

        for row in range(2, max_row + 1):
            formula = f'=TRIM(LOWER({status_letter}{row}))'
            self.worksheet.cell(row=row, column=self.formula_col).value = formula

        # Oculta a coluna da fórmula
        self.worksheet.column_dimensions[get_column_letter(self.formula_col)].hidden = True

    def aplicar_formatacao_condicional(self):
        from openpyxl.styles import PatternFill
        from openpyxl.styles.differential import DifferentialStyle
        from openpyxl.formatting import Rule

        max_col = self.status_col
        max_row = self.worksheet.max_row
        status_letter = get_column_letter(self.status_col)
        faixa = f"A2:{get_column_letter(max_col)}{max_row}"  # intervalo completo

        regras = [
            {"valor": "s", "cor": "FF1493"},  # rosa forte
            {"valor": "t", "cor": "00FF00"},  # verde forte
            {"valor": "d", "cor": "FF0000"},  # vermelho forte
        ]

        for regra in regras:
            formula = f'TRIM(LOWER(${status_letter}2))="{regra["valor"]}"'  # usa linha 2 como base (relativa)
            dxf = DifferentialStyle(
                fill=PatternFill(
                    start_color=regra["cor"],
                    end_color=regra["cor"],
                    fill_type='solid'
                )
            )
            rule = Rule(type="expression", dxf=dxf, stopIfTrue=True, formula=[formula])
            self.worksheet.conditional_formatting.add(faixa, rule)

    def salvar(self):
        dirname, filename = os.path.split(self.filepath)
        nome_base, _ = os.path.splitext(filename)
        novo_nome = f"{nome_base}_atualizado.xlsx"
        novo_caminho = os.path.join(dirname, novo_nome)
        self.workbook.save(novo_caminho)
        return novo_caminho

    def processar(self):
        self.adicionar_colunas_status_funcao()
        self.aplicar_formatacao_condicional()
        return self.salvar()